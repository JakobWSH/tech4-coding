import pyomo.environ as pyo
from pathlib import Path
from openpyxl import load_workbook  # reads precise cell ranges

solver = "appsi_highs"
SOLVER = pyo.SolverFactory(solver)
assert SOLVER.available(), "HiGHS (appsi_highs) not available"

xlsx = Path(__file__).with_name("05c_Southwestern_Airways.xlsx")
wb = load_workbook(xlsx, data_only=True)
ws = wb["Southwestern"]

# --- read ranges that exist in your file ---
# costs: C5:N5  (12 sequences)
cost_row = [c.value for c in ws["C5":"N5"][0]]
# flights: B8:B18 (names)
flights = [r[0].value for r in ws["B8":"B18"]]
# incidence: C8:N18  (11 flights × 12 sequences)
inc = [[c.value for c in row] for row in ws["C8":"N18"]]

J = list(range(1, len(cost_row) + 1))   # sequences 1..12
I = flights                              # flight labels from the sheet

# build dicts
cost = {j: float(cost_row[j-1]) for j in J}
A = {(I[i], J[j]): int(inc[i][j]) for i in range(len(I)) for j in range(len(J))}

# --- model: choose exactly 3 sequences, cover every flight, min cost ---
m = pyo.ConcreteModel("SW_Airways")
m.I = pyo.Set(initialize=I)
m.J = pyo.Set(initialize=J)
m.x = pyo.Var(m.J, domain=pyo.Binary)  # 0/1 variables (Pyomo Binary domain)

m.num_crews = pyo.Constraint(expr=sum(m.x[j] for j in m.J) == 3)
m.cover = pyo.Constraint(m.I, rule=lambda m, i: sum(A[i, j]*m.x[j] for j in m.J) >= 1)
m.obj = pyo.Objective(expr=sum(cost[j]*m.x[j] for j in m.J), sense=pyo.minimize)

SOLVER.solve(m)

chosen = [j for j in m.J if pyo.value(m.x[j]) > 0.5]
print("Chosen sequences:", chosen)
print("Total cost:", pyo.value(m.obj))