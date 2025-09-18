import pyomo.environ as pyo
from pathlib import Path
from openpyxl import load_workbook

solver = "appsi_highs"
opt = pyo.SolverFactory(solver)
assert opt.available(), "HiGHS (appsi_highs) not available"

# --- Excel next to this script ---
here = Path(__file__).parent
xlsx = here / "06b_P&T_REV.xlsx"
if not xlsx.exists():
    raise FileNotFoundError(f"Expected Excel here:\n  {xlsx}")

# Read evaluated values (fast, read-only)
wb = load_workbook(xlsx, data_only=True, read_only=True)  # values_only via iter_rows below
ws = wb["P&T"]

# ---- RANGES (exactly as in your sheet) ----
# Destinations (warehouse headers): D3:G3
D = list(next(ws.iter_rows(min_row=3, max_row=3, min_col=4, max_col=7, values_only=True)))

# Sources (canneries): C5:C7
S = [r[0] for r in ws.iter_rows(min_row=5, max_row=7, min_col=3, max_col=3, values_only=True)]

# Unit shipping costs: D5:G7 (rows align with S, cols with D)
unit_cost = {}
for i, row in enumerate(ws.iter_rows(min_row=5, max_row=7, min_col=4, max_col=7, values_only=True)):
    for j, val in enumerate(row):
        unit_cost[(S[i], D[j])] = float(val or 0.0)

# Original / maximum supply: K5:K7
orig_supply = {}
for i, row in enumerate(ws.iter_rows(min_row=5, max_row=7, min_col=11, max_col=11, values_only=True)):
    orig_supply[S[i]] = float(row[0] or 0.0)

# Demands: D17:G17
demand = {}
for j, val in enumerate(next(ws.iter_rows(min_row=17, max_row=17, min_col=4, max_col=7, values_only=True))):
    demand[D[j]] = float(val or 0.0)

FIX = 5000.0  # fixed cost per cannery used (given in sheet/lecture)

# Quick feasibility sanity check (avoid solver confusion)
if sum(2.0 * orig_supply[s] for s in S) < sum(demand[d] for d in D):
    raise ValueError("Infeasible data: 2× total supply < total demand. Check ranges/assumptions.")

# ---- Model ----
m = pyo.ConcreteModel("PT_fixed_charge")
m.S = pyo.Set(initialize=S)
m.D = pyo.Set(initialize=D)

m.x = pyo.Var(m.S, m.D, domain=pyo.NonNegativeReals)  # flows
m.y = pyo.Var(m.S, domain=pyo.Binary)                  # open cannery?

# Supply if opened: sum_d x[s,d] ≤ 2 × orig_supply[s] * y[s]
def supply_rule(m, s):
    return sum(m.x[s, d] for d in m.D) <= 2.0 * orig_supply[s] * m.y[s]
m.supply = pyo.Constraint(m.S, rule=supply_rule)

# Meet each demand
def demand_rule(m, d):
    return sum(m.x[s, d] for s in m.S) >= demand[d]
m.demand = pyo.Constraint(m.D, rule=demand_rule)

# Minimize shipping + fixed charges
m.obj = pyo.Objective(
    expr=sum(unit_cost[s, d] * m.x[s, d] for s in m.S for d in m.D)
        + sum(FIX * m.y[s] for s in m.S),
    sense=pyo.minimize
)

opt.solve(m)

# ---- Report ----
used = {s: int(pyo.value(m.y[s])) for s in m.S}
print("Used canneries:", used)
print("Total cost:", round(pyo.value(m.obj), 2))
for s in m.S:
    flow = {d: pyo.value(m.x[s, d]) for d in m.D if pyo.value(m.x[s, d]) > 1e-8}
    if flow:
        print(f"{s} -> {flow}")