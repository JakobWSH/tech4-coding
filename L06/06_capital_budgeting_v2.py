import pyomo.environ as pyo
from pathlib import Path
from openpyxl import load_workbook

solver = "appsi_highs"
opt = pyo.SolverFactory(solver)
assert opt.available(), "HiGHS (appsi_highs) not available"

# Excel next to this script
here = Path(__file__).parent
xlsx = here / "06c_CapitalBudgeting_v2.xlsx"
if not xlsx.exists():
    raise FileNotFoundError(f"Expected Excel here:\n  {xlsx}")

wb = load_workbook(xlsx, data_only=True, read_only=True)
ws = wb["Modified"]

# Ranges per sheet
projects = list(next(ws.iter_rows(min_row=4, max_row=4, min_col=3, max_col=7, values_only=True)))
npv_row  = list(next(ws.iter_rows(min_row=5, max_row=5, min_col=3, max_col=7, values_only=True)))
i0_row   = list(next(ws.iter_rows(min_row=6, max_row=6, min_col=3, max_col=7, values_only=True)))

# Bonuses and budget
bonus12  = float(next(ws.iter_rows(min_row=7, max_row=7, min_col=3, max_col=3, values_only=True))[0] or 0.0)  # C7
bonus45  = float(next(ws.iter_rows(min_row=7, max_row=7, min_col=6, max_col=6, values_only=True))[0] or 0.0)  # F7
budget_cell = next(ws.iter_rows(min_row=13, max_row=13, min_col=10, max_col=10, values_only=True))[0]        # J13
budget = float(budget_cell) if budget_cell is not None else 20.0  # default if empty

NPV = {projects[j]: float(npv_row[j] or 0.0) for j in range(len(projects))}
I0  = {projects[j]: float(i0_row[j]  or 0.0) for j in range(len(projects))}

# Model
m = pyo.ConcreteModel("CapBudget_bonus")
m.J = pyo.Set(initialize=projects)
m.x = pyo.Var(m.J, domain=pyo.Binary)  # 0/1 choose project  # Pyomo Binary domain
# two auxiliary binaries that "turn on" only if both projects in the pair are chosen
m.z12 = pyo.Var(domain=pyo.Binary)
m.z45 = pyo.Var(domain=pyo.Binary)

# Budget
m.budget = pyo.Constraint(expr=sum(I0[j]*m.x[j] for j in m.J) <= budget)

# Logical AND linearization (standard 3 inequalities):
# z <= x, z <= y, z >= x + y - 1
p1, p2, p4, p5 = projects[0], projects[1], projects[3], projects[4]
m.c12a = pyo.Constraint(expr=m.z12 <= m.x[p1])
m.c12b = pyo.Constraint(expr=m.z12 <= m.x[p2])
m.c12c = pyo.Constraint(expr=m.z12 >= m.x[p1] + m.x[p2] - 1)

m.c45a = pyo.Constraint(expr=m.z45 <= m.x[p4])
m.c45b = pyo.Constraint(expr=m.z45 <= m.x[p5])
m.c45c = pyo.Constraint(expr=m.z45 >= m.x[p4] + m.x[p5] - 1)

# Objective: base NPV + bonuses if both projects in a pair are selected
m.obj = pyo.Objective(
    expr=sum(NPV[j]*m.x[j] for j in m.J) + bonus12*m.z12 + bonus45*m.z45,
    sense=pyo.maximize
)

opt.solve(m)

print("Chosen:", {j: int(pyo.value(m.x[j])) for j in m.J})
print("Bonus flags (1&2, 4&5):", int(pyo.value(m.z12)), int(pyo.value(m.z45)))
print("Total NPV:", round(pyo.value(m.obj), 4))
print("Budget used:", round(sum(I0[j]*pyo.value(m.x[j]) for j in m.J), 4), "of", budget)