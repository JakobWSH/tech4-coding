import pyomo.environ as pyo
from pathlib import Path
from openpyxl import load_workbook

solver = "appsi_highs"
opt = pyo.SolverFactory(solver)
assert opt.available(), "HiGHS (appsi_highs) not available"

# Excel next to this script
here = Path(__file__).parent
xlsx = here / "06c_CapitalBudgeting_v1.xlsx"
if not xlsx.exists():
    raise FileNotFoundError(f"Expected Excel here:\n  {xlsx}")

wb = load_workbook(xlsx, data_only=True, read_only=True)
ws = wb["Original"]

# Ranges (exactly as in the sheet)
projects = list(next(ws.iter_rows(min_row=4, max_row=4, min_col=3, max_col=7, values_only=True)))
npv_row  = list(next(ws.iter_rows(min_row=5, max_row=5, min_col=3, max_col=7, values_only=True)))
i0_row   = list(next(ws.iter_rows(min_row=6, max_row=6, min_col=3, max_col=7, values_only=True)))
# Budget in J13
budget   = float(next(ws.iter_rows(min_row=13, max_row=13, min_col=10, max_col=10, values_only=True))[0])

NPV = {projects[j]: float(npv_row[j] or 0.0) for j in range(len(projects))}
I0  = {projects[j]: float(i0_row[j]  or 0.0) for j in range(len(projects))}

# Model
m = pyo.ConcreteModel("CapBudget_knapsack")
m.J = pyo.Set(initialize=projects)
m.x = pyo.Var(m.J, domain=pyo.Binary)  # binary decision per project (Pyomo Binary domain)  # docs: domain=pyo.Binary

# Budget and objective
m.budget = pyo.Constraint(expr=sum(I0[j]*m.x[j] for j in m.J) <= budget)
m.obj    = pyo.Objective(expr=sum(NPV[j]*m.x[j] for j in m.J), sense=pyo.maximize)

opt.solve(m)

print("Chosen:", {j: int(pyo.value(m.x[j])) for j in m.J})
print("Total NPV:", round(pyo.value(m.obj), 4))
print("Budget used:", round(sum(I0[j]*pyo.value(m.x[j]) for j in m.J), 4), "of", budget)