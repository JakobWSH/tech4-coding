import pyomo.environ as pyo
from pathlib import Path
from openpyxl import load_workbook

solver = "appsi_highs"
SOLVER = pyo.SolverFactory(solver)
assert SOLVER.available(), "HiGHS (appsi_highs) not available"

# ---- read ranges exactly like the workbook ----
here = Path(__file__).parent
xlsx = here / "06a_AirPollution_Nori_and_Leets_REV.xlsx"
if not xlsx.exists():
    raise FileNotFoundError(f"Expected Excel here: {xlsx}")

# Read evaluated values efficiently (no formula evaluation)
wb = load_workbook(xlsx, data_only=True, read_only=True)   # openpyxl docs on flags
ws = wb["Nori and Leets"]

# Costs for 6 method-at-furnace options: stacks/filter/fuel × (blast, open-hearth)
# Row 6 holds variable costs for each option (C6:H6)
var_cost = [float((c.value or 0)) for c in ws["C6":"H6"][0]]

# Row 8 has fixed costs only for smokestacks (C8:D8); others are 0
fix_costs = [float((c.value or 0)) for c in ws["C8":"D8"][0]]  # e.g., [2.0, 2.0]

# Pollutant names & required minimum reductions (B12:B14 and K12:K14)
pollutants = [r[0].value for r in ws["B12":"B14"]]
min_reduction = {p: float(cell[0].value) for p, cell in zip(pollutants, ws["K12":"K14"])}

# Reduction matrix (millions of lbs) if an option is used at 100%
# C12:H14  (rows: pollutants, cols: 6 options)
red = {
    (pollutants[i], j + 1): float((cell.value or 0))
    for i, row in enumerate(ws["C12":"H14"])
    for j, cell in enumerate(row)
}

# Upper bounds on “fraction used” for each option (C21:H21) — read cell values
ub = {j + 1: float((cell.value if cell.value is not None else 1.0))
      for j, cell in enumerate(ws["C21":"H21"][0])}

# ---- model ----
m = pyo.ConcreteModel("Nori_Leets_FixedCharge")
J = range(1, 7)  # 6 options in the same order as the sheet

# fraction used (0..ub[j]); Pyomo domains use `domain=...`
m.x = pyo.Var(J, domain=pyo.NonNegativeReals)
# fixed-charge binaries: stacks at Blast(1), Open-Hearth(2)
m.y = pyo.Var([1, 2], domain=pyo.Binary)

# capacity bounds on fractions
m.ub = pyo.Constraint(J, rule=lambda m, j: m.x[j] <= ub[j])

# link stacks to binaries via Big-M (M = ub[j])
m.bigM1 = pyo.Constraint(expr=m.x[1] <= ub[1] * m.y[1])
m.bigM2 = pyo.Constraint(expr=m.x[2] <= ub[2] * m.y[2])
# (filters & fuels have no fixed charge → no binary needed)

# meet pollutant reduction minimums
def meet_rule(m, p):
    return sum(red[p, j] * m.x[j] for j in J) >= min_reduction[p]
m.meet = pyo.Constraint(pollutants, rule=meet_rule)

# objective: variable cost + fixed costs for stacks (taken from the sheet)
var_term = sum(var_cost[j - 1] * m.x[j] for j in J)
fix_term = fix_costs[0] * m.y[1] + fix_costs[1] * m.y[2]
m.obj = pyo.Objective(expr=var_term + fix_term, sense=pyo.minimize)

SOLVER.solve(m)

print("Fractions used:", {j: round(pyo.value(m.x[j]), 4) for j in J})
print("Stacks used (Blast, Open-Hearth):", int(pyo.value(m.y[1])), int(pyo.value(m.y[2])))
print("Total cost ($M):", round(pyo.value(m.obj), 4))